"""Deep analysis of the recruitment tracker Excel file."""
import json
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILE = "Recruitment_Staffing_Tracker_RADiiX_INFINITEii_Year_2026_SYNCED.xlsx"

wb = load_workbook(FILE, data_only=True)
print("=== SHEETS ===")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

    # Print rows 1-4 (header structure)
    print("\n--- HEADER ROWS 1-4 ---")
    for r in range(1, 5):
        cells = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is not None and str(val).strip():
                cells.append(f"{get_column_letter(c)}={repr(val)[:80]}")
        if cells:
            print(f"Row {r}: " + " | ".join(cells[:20]))
            if len(cells) > 20:
                print(f"  ... +{len(cells)-20} more")

    # Column headers from row 4
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(4, c).value
        if h:
            headers[c] = str(h).strip()

    print(f"\n--- COLUMN HEADERS (row 4): {len(headers)} columns ---")
    for c, h in headers.items():
        print(f"  {get_column_letter(c)} ({c}): {h}")

    # Data analysis from row 5
    data_start = 5
    total_data_rows = 0
    non_empty_position_rows = 0
    col_stats = {}

    for c, h in headers.items():
        col_stats[c] = {
            "letter": get_column_letter(c),
            "header": h,
            "filled": 0,
            "empty": 0,
            "samples": [],
            "unique_values": set(),
        }

    for r in range(data_start, ws.max_row + 1):
        row_has_data = False
        for c in headers:
            val = ws.cell(r, c).value
            if val is not None and str(val).strip() != "":
                row_has_data = True
                col_stats[c]["filled"] += 1
                s = str(val).strip()
                if len(col_stats[c]["samples"]) < 5:
                    col_stats[c]["samples"].append(s[:100])
                if len(col_stats[c]["unique_values"]) < 30:
                    col_stats[c]["unique_values"].add(s[:80])
            else:
                col_stats[c]["empty"] += 1
        if row_has_data:
            total_data_rows += 1
        pos_col = next((c for c, h in headers.items() if "position" in h.lower() and "receiving" not in h.lower()), None)
        if pos_col:
            pos_val = ws.cell(r, pos_col).value
            if pos_val and str(pos_val).strip():
                non_empty_position_rows += 1

    print(f"\n--- DATA ROW STATS (from row {data_start}) ---")
    print(f"Rows with any data: {total_data_rows}")
    print(f"Rows with position filled: {non_empty_position_rows}")

    print("\n--- PER-COLUMN FILL RATE (columns with data) ---")
    filled_cols = []
    for c in sorted(col_stats.keys()):
        st = col_stats[c]
        if st["filled"] > 0:
            filled_cols.append(st)
            pct = 100 * st["filled"] / max(total_data_rows, 1)
            print(f"\n  [{st['letter']}] {st['header']}")
            print(f"    Filled: {st['filled']}/{total_data_rows} rows ({pct:.0f}%)")
            print(f"    Samples: {st['samples'][:3]}")
            uniq = list(st["unique_values"])[:15]
            if len(uniq) <= 15:
                print(f"    Unique ({len(st['unique_values'])}): {uniq}")
            else:
                print(f"    Unique count: {len(st['unique_values'])}")

    # Show completely empty columns (that have headers)
    empty_header_cols = [st for c, st in col_stats.items() if st["filled"] == 0]
    if empty_header_cols:
        print(f"\n--- COLUMNS WITH HEADERS BUT NO DATA ({len(empty_header_cols)}) ---")
        for st in empty_header_cols:
            print(f"  [{st['letter']}] {st['header']}")

    # Sample full rows with most data
    print("\n--- SAMPLE DATA ROWS (first 3 with position) ---")
    pos_col = next((c for c, h in headers.items() if h == "Position Name / Role"), None)
    if not pos_col:
        pos_col = next((c for c, h in headers.items() if "position" in h.lower()), None)
    shown = 0
    for r in range(data_start, ws.max_row + 1):
        if pos_col and ws.cell(r, pos_col).value:
            print(f"\nRow {r}:")
            for c, h in headers.items():
                val = ws.cell(r, c).value
                if val is not None and str(val).strip():
                    print(f"  {get_column_letter(c)} ({h}): {repr(val)[:120]}")
            shown += 1
            if shown >= 3:
                break
